from openpyxl import Workbook

wb = Workbook()

ws = wb.active
ws.title = "Nadosheet"

ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"])
print(ws["A1"].value) #값을 찍을 때 이렇게
print(ws["A10"].value) # 값이 없을 땐 None으로 출력
print(ws.cell(row=1,column=1).value) #좌표 값으로 값을 찾을 수도 있음 , column이 A,B,C,,, 

c = ws.cell(column=3, row=1, value=10) # 바로 value를 지정
print(c.value)
index=1
from random import *
for x in range(1,11):
  for y in range(1,11):
    # ws.cell(row=x, column=y, value=randint(0, 100))
    ws.cell(row=x, column=y, value=index)
    index+=1

wb.save("sample.xlsx")