from openpyxl import Workbook
wb = Workbook() # 새 워크북 생성

ws = wb.create_sheet() # 새로운 시트를 기본이름으로 생성
ws.title = "Mysheet" # 시트 이름 변경
ws.sheet_properties.tabColor = "ff66ff" # 탭 생상 변경, #없이 넣어야함

ws1 = wb.create_sheet("YourSheet") # 주어진 이름으로 sheet 생성

ws2 = wb.create_sheet("NewSheet", 2) #2번째 인덱스에 시트 생성  

new_ws = wb["NewSheet"] # 딕셔너리 형태로 시트에 접근이 가능

print(wb.sheetnames) #모든 시트 이름 확인

# Sheet 복사
new_ws["A1"] = "Test" # A1 cell에 test를 입력하는 행위
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")
