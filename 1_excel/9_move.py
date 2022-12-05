from openpyxl import load_workbook

wb = load_workbook("sample.xlsx")
ws = wb.active

ws.move_range("B1:C11", rows=0, cols=1)
ws["B1"].value = "국어"

ws.move_range("C1:C11", rows=5, cols=-1) #원래 데이터를 덮어쓰게 됨


wb.save("sample_korean.xlsx")