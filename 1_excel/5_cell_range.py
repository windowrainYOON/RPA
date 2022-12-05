from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active

#1줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"])
for i in range(1, 11):
  ws.append([i, randint(0, 100), randint(0, 100)])

col_B = ws["B"] #B 의 데이터를 다 가져옴
for cell in col_B:
  print(cell.value)

col_range = ws["B:C"] #B,C 열을 함께 가져오기

for cols in col_range:
  for cell in cols:
    print(cell.value)

#====================
row_title = ws[1] #1번째 행 가져오기
for cell in row_title:
  print(cell.value)

row_range = ws[2:6] # 2,3,4,5,6 번 행을 가져오기
for rows in row_range:
  for cell in rows:
    print(cell.value, end=" ")
  print()

#===================
from openpyxl.utils.cell import coordinate_from_string

row_range_max = ws[2:ws.max_row] #모든 데이터를 다 받아오기
for rows in row_range_max:
  for cell in rows:
    print(cell.value, end=" ")
    print(cell.coordinate, end=" ") #cell의 좌표 정도를 가져올 수 있음
    xy = coordinate_from_string(cell.coordinate) #('A', 5)
    print(xy, end=" ")
    print(xy[0], end="")
    print(xy[1], end=" ")
  print()

#==================
#전체 row
print(tuple(ws.rows))
for row in tuple(ws.rows): #한 row에 있는 모든 값을 가져옴
  print(row[1].value)

print(tuple(ws.columns))
for column in tuple(ws.columns):
  print(column[0].value)

for row in ws.iter_rows():
  print(row[1].value)

for column in ws.iter_cols():
  print(column[0].value)

#====================
for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3):
  print(row[0].value, row[1].value)

wb.save("sample.xlsx")