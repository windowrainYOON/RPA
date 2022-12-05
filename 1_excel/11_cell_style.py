from openpyxl.styles import Font, Border, Side
from openpyxl import load_workbook

wb = load_workbook("sample.xlsx")
ws = wb.active

a1 = ws["A1"]
b1 = ws["B1"]
c1 = ws["C1"]

ws.column_dimensions["A"].width = 5 # A열의 너비를 5로 설정
ws.row_dimensions[1].height = 50 #첫번째 줄의 높이를 50으로

a1.font = Font(color="FF0000", italic=True, bold=True) #글자 색상, 스타일 적용
b1.font = Font(color="CC33FF", name="Arial", strike=True)
c1.font = Font(color="0000FF", size=20, underline="single")

thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border
wb.save("sample_style.xlsx")